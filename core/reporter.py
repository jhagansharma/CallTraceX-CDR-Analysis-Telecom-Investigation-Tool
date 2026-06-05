"""
CDR Reporter v3.1 - Law Enforcement Edition (Production Build)
Generates professional Excel report with i9/JAS-style Smart Report
Single Excel file with Smart Report as first sheet + all analysis sheets
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ============================================================
# STYLES
# ============================================================
DARK_BLUE = PatternFill('solid', start_color='1F4E79')
MED_BLUE = PatternFill('solid', start_color='2E75B6')
LIGHT_BLUE = PatternFill('solid', start_color='5B9BD5')
LIGHT_GRAY = PatternFill('solid', start_color='F2F2F2')
WHITE_FILL = PatternFill('solid', start_color='FFFFFF')
SEP_FILL = PatternFill('solid', start_color='D6E4F0')  # light blue separator
ZEBRA_FILL = PatternFill('solid', start_color='EBF1F8')  # very light blue for alternating rows
ACCENT_FILL = PatternFill('solid', start_color='DAEEF3')  # light teal accent

WHITE_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
WHITE_SM = Font(name='Arial', bold=True, color='FFFFFF', size=9)
TITLE_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=14)
SUBTITLE_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
LABEL_FONT = Font(name='Arial', bold=True, size=9, color='1F4E79')
VALUE_FONT = Font(name='Arial', size=9)
VALUE_BOLD = Font(name='Arial', bold=True, size=9)
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
STAT_NUM_FONT = Font(name='Arial', bold=True, size=12, color='1F4E79')
FOOTER_FONT = Font(name='Arial', italic=True, size=8, color='7D8590')
SR_FONT = Font(name='Arial', bold=True, size=9, color='4472C4')  # serial number font

THIN = Border(
    left=Side('thin'), right=Side('thin'),
    top=Side('thin'), bottom=Side('thin')
)
LIGHT_BORDER = Border(
    left=Side('thin', color='D3D3D3'), right=Side('thin', color='D3D3D3'),
    top=Side('thin', color='D3D3D3'), bottom=Side('thin', color='D3D3D3')
)
BOTTOM_THICK = Border(bottom=Side('medium', color='1F4E79'))
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(vertical='center', wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')


def _cell(ws, r, c, val, font=VALUE_FONT, fill=None, align=None, border=THIN):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    return cell


def _header_row(ws, row, headers, fill=LIGHT_BLUE):
    for i, h in enumerate(headers, 1):
        _cell(ws, row, i, h, WHITE_SM, fill, CENTER)


class ExcelReporter:
    """Generate formatted Excel report with i9/JAS-style Smart Report"""

    def __init__(self, target_number):
        self.target = target_number

    def generate_report(self, analysis_results, output_path, cleaned_df, metadata):
        """
        Main entry point. Generates complete Excel report.

        Args:
            analysis_results: dict of {sheet_name: DataFrame} from CDRAnalyzer
            output_path: where to save .xlsx
            cleaned_df: the cleaned DataFrame from CDRParser
            metadata: dict from parser.get_metadata_dict()

        Returns:
            output_path
        """
        logger.info("\n" + "=" * 80)
        logger.info("GENERATING EXCEL REPORT")
        logger.info("=" * 80)

        wb = Workbook()
        wb.remove(wb.active)

        # SHEET 1: Professional Smart Report
        logger.info("  Building Smart Report...")
        self._create_smart_report(wb, cleaned_df, metadata, analysis_results)

        # REMAINING SHEETS: All analysis results
        for sheet_name, df in analysis_results.items():
            if sheet_name == 'Smart Report':
                continue
            try:
                if df is None or len(df) == 0:
                    logger.info(f"  Skipping: {sheet_name} (empty)")
                    continue
            except Exception:
                logger.info(f"  Skipping: {sheet_name} (invalid data)")
                continue

            logger.info(f"  Writing: {sheet_name} ({len(df)} rows)")
            try:
                ws = wb.create_sheet(sheet_name)

                # Title row
                max_col = min(len(df.columns), 20)
                if max_col > 1:
                    from openpyxl.utils import get_column_letter
                    end_col = get_column_letter(max_col)
                    ws.merge_cells(f'A1:{end_col}1')
                title_text = self._sheet_title(sheet_name)
                _cell(ws, 1, 1, f'{title_text} : {self.target}', TITLE_FONT, DARK_BLUE, CENTER)

                # Header row
                col_names = df.columns.tolist()
                _header_row(ws, 2, col_names)

                # Identify duration columns for formatting
                dur_cols = {i for i, c in enumerate(col_names)
                            if 'DURATION' in str(c).upper() or c == 'DUR'}

                # Data rows with alternating colors
                for ridx, (_, row) in enumerate(df.iterrows(), 3):
                    fill = ZEBRA_FILL if (ridx - 3) % 2 == 0 else WHITE_FILL
                    for cidx, val in enumerate(row, 1):
                        display = self._format_value(val)
                        # Format duration columns as HH:MM:SS
                        if cidx - 1 in dur_cols and display != '':
                            try:
                                sec = int(float(display))
                                if sec > 0:
                                    h, rem = divmod(sec, 3600)
                                    m, s = divmod(rem, 60)
                                    display = f"{h:02d}:{m:02d}:{s:02d}" if h > 0 else f"{m:02d}:{s:02d}"
                                else:
                                    display = ''
                            except (ValueError, TypeError):
                                pass
                        _cell(ws, ridx, cidx, display, VALUE_FONT, fill, border=LIGHT_BORDER)

                ws.freeze_panes = 'A3'

                # Print settings for data sheets
                ws.sheet_properties.pageSetUpPr = None
                ws.page_setup.orientation = 'landscape'
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0
                ws.print_title_rows = '1:2'
            except Exception as e:
                logger.warning(f"  ERROR writing {sheet_name}: {e}")
                continue

        # RAW CDR (append raw file content)
        logger.info("  Writing Raw CDR...")
        self._add_raw_cdr(wb, metadata)

        # Format all sheets
        self._format_all(wb)

        # Save
        try:
            wb.save(output_path)
        except PermissionError:
            raise PermissionError(
                f"Cannot save report - file is open in Excel!\n\n"
                f"Please close the file and try again:\n{output_path}"
            )
        except Exception as e:
            raise RuntimeError(f"Failed to save report: {e}")
        logger.info(f"\n  Report saved: {output_path}")
        logger.info(f"  Sheets: {len(wb.sheetnames)}")
        return output_path

    # ==============================================================
    # SMART REPORT (Sheet 1) - i9/JAS format - PROFESSIONAL LAYOUT
    # ==============================================================

    @staticmethod
    def _safe(val):
        """Return empty string for None/nan/NaT values"""
        try:
            if val is None:
                return ''
            s = str(val)
            if s.lower() in ('nan', 'nat', 'none', '', '-'):
                return ''
            return val
        except Exception:
            return ''

    @staticmethod
    def _safe_date(val, fmt='%d-%b-%Y'):
        """Safely format any value as date string, never crashes"""
        try:
            if val is None:
                return ''
            if hasattr(val, 'strftime') and pd.notna(val):
                return val.strftime(fmt)
            dt = pd.to_datetime(val)
            if pd.notna(dt):
                return dt.strftime(fmt)
            return ''
        except Exception:
            return '' 

    @staticmethod
    def _fmt_dur(seconds):
        """Format seconds as HH:MM:SS"""
        try:
            sec = int(seconds)
            if sec <= 0:
                return ''
            h, remainder = divmod(sec, 3600)
            m, s = divmod(remainder, 60)
            if h > 0:
                return f"{h:02d}:{m:02d}:{s:02d}"
            return f"{m:02d}:{s:02d}"
        except (ValueError, TypeError):
            return ''

    def _section_header(self, ws, row, title, col_start=1, col_end=10):
        """Create a full-width section header with dark blue background"""
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        _cell(ws, row, col_start, title, SUBTITLE_FONT, MED_BLUE, CENTER)
        ws.row_dimensions[row].height = 24

    @staticmethod
    def _row_fill(idx):
        """Return alternating fill for zebra striping"""
        return ZEBRA_FILL if idx % 2 == 0 else WHITE_FILL

    def _sub_header_pair(self, ws, row, left_title, right_title):
        """Create paired sub-headers for left (A-D) and right (F-J) tables"""
        ws.merge_cells(f'A{row}:D{row}')
        _cell(ws, row, 1, left_title, WHITE_SM, LIGHT_BLUE, CENTER)
        _cell(ws, row, 5, '', VALUE_FONT, SEP_FILL, border=None)
        ws.merge_cells(f'F{row}:J{row}')
        _cell(ws, row, 6, right_title, WHITE_SM, LIGHT_BLUE, CENTER)
        ws.row_dimensions[row].height = 20

    def _sep_col(self, ws, row):
        """Style the separator column E for a given row"""
        _cell(ws, row, 5, '', VALUE_FONT, SEP_FILL, border=None)

    def _create_smart_report(self, wb, df, meta, results):
        try:
            self._build_smart_report(wb, df, meta, results)
        except Exception as e:
            logger.warning(f'  Smart Report error: {e}')
            import traceback; logger.warning(traceback.format_exc())

    def _build_smart_report(self, wb, df, meta, results):
        ws = wb.create_sheet('Smart Report', 0)
        row = 1

        # --- TITLE BAR ---
        ws.merge_cells(f'A{row}:J{row}')
        _cell(ws, row, 1, 'SMART REPORT - CDR FORENSIC ANALYSIS', TITLE_FONT, DARK_BLUE, CENTER)
        ws.row_dimensions[row].height = 32
        row += 2

        # --- BASIC INFO BLOCK ---
        target = self._safe(meta.get('input_value', 'Unknown'))
        date_range = self._safe(meta.get('date_range', 'N/A'))
        try:
            parts = date_range.split(' to ')
            if len(parts) == 2:
                d1 = pd.to_datetime(parts[0].strip()).strftime('%d-%b-%Y')
                d2 = pd.to_datetime(parts[1].strip()).strftime('%d-%b-%Y')
                date_range = f"{d1}  TO  {d2}"
        except Exception:
            pass

        # Detect operator
        operator = 'Unknown'
        if len(df) > 0 and 'OP_LABEL' in df.columns:
            ops = df[df['OP_LABEL'] != 'Service SMS']['OP_LABEL']
            if len(ops) > 0 and len(ops.mode()) > 0:
                operator = self._safe(ops.mode().iloc[0])

        # Calculate total calls/duration for display
        total_records = len(df)
        total_dur = int(df['DUR_SECONDS'].sum()) if 'DUR_SECONDS' in df.columns else 0

        left_info = [
            ('Target No.', target),
            ('CDR Period', date_range),
            ('Operator', operator),
            ('Circle', self._safe(meta.get('circle', ''))),
            ('Total Records', str(total_records)),
        ]
        right_info = [
            ('Target Name', self._safe(meta.get('subscriber_name', ''))),
            ('Father/Husband', self._safe(meta.get('father_husband_name', ''))),
            ('Ticket No.', self._safe(meta.get('ticket_number', ''))),
            ('SIM Activation', self._safe(meta.get('sim_activation_date', ''))),
            ('Total Duration', self._fmt_dur(total_dur)),
        ]

        for (ll, lv), (rl, rv) in zip(left_info, right_info):
            _cell(ws, row, 1, ll, LABEL_FONT, LIGHT_GRAY, LEFT_WRAP, THIN)
            ws.merge_cells(f'B{row}:D{row}')
            _cell(ws, row, 2, lv, VALUE_BOLD, WHITE_FILL, LEFT_WRAP, THIN)
            self._sep_col(ws, row)
            _cell(ws, row, 6, rl, LABEL_FONT, LIGHT_GRAY, LEFT_WRAP, THIN)
            ws.merge_cells(f'G{row}:J{row}')
            _cell(ws, row, 7, rv, VALUE_BOLD, WHITE_FILL, LEFT_WRAP, THIN)
            ws.row_dimensions[row].height = 18
            row += 1
        row += 1

        # --- CALL SUMMARY STATISTICS ---
        self._section_header(ws, row, 'CALL SUMMARY')
        row += 1

        out_call = len(df[df['CALL_TYPE_LABEL'] == 'OUT-CALL']) if 'CALL_TYPE_LABEL' in df.columns else 0
        in_call = len(df[df['CALL_TYPE_LABEL'] == 'IN-CALL']) if 'CALL_TYPE_LABEL' in df.columns else 0
        out_sms = len(df[df['CALL_TYPE_LABEL'] == 'OUT-SMS']) if 'CALL_TYPE_LABEL' in df.columns else 0
        in_sms = len(df[df['CALL_TYPE_LABEL'] == 'IN-SMS']) if 'CALL_TYPE_LABEL' in df.columns else 0
        fwd_call = len(df[df['CALL_TYPE_LABEL'] == 'FWD-CALL']) if 'CALL_TYPE_LABEL' in df.columns else 0
        bsm_count = len(df[df['CALL_TYPE_LABEL'].str.contains('BSM', na=False)]) if 'CALL_TYPE_LABEL' in df.columns else 0
        unique_contacts = 0
        if 'OTHER_PARTY' in df.columns:
            valid_contacts = df['OTHER_PARTY'].astype(str).str.strip()
            valid_contacts = valid_contacts[~valid_contacts.isin(['', 'nan', 'None', '-'])]
            unique_contacts = valid_contacts.nunique()
        unique_imei = 0
        if 'IMEI' in df.columns:
            valid_imei = df['IMEI'].astype(str).str.strip()
            valid_imei = valid_imei[~valid_imei.isin(['', 'nan', 'None', '-'])]
            unique_imei = valid_imei.nunique()
        unique_cells = 0
        if 'FIRST_CELL_ID' in df.columns:
            valid_cells = df['FIRST_CELL_ID'].astype(str).str.strip()
            valid_cells = valid_cells[~valid_cells.isin(['', 'nan', 'None', '-', '---'])]
            unique_cells = valid_cells.nunique()

        stat_left = [('OUT-CALL', out_call), ('IN-CALL', in_call), ('OUT-SMS', out_sms), ('IN-SMS', in_sms)]
        stat_right = [('FWD-CALL', fwd_call), ('CONTACTS', unique_contacts), ('IMEI', unique_imei), ('LOCATIONS', unique_cells)]

        # Left stat headers (A-D)
        for i, (h, _) in enumerate(stat_left, 1):
            _cell(ws, row, i, h, WHITE_SM, LIGHT_BLUE, CENTER)
        self._sep_col(ws, row)
        # Right stat headers (F-I)
        for i, (h, _) in enumerate(stat_right, 6):
            _cell(ws, row, i, h, WHITE_SM, LIGHT_BLUE, CENTER)
        _cell(ws, row, 10, 'TOTAL', WHITE_SM, LIGHT_BLUE, CENTER)
        ws.row_dimensions[row].height = 20
        row += 1

        # Left stat values
        for i, (_, v) in enumerate(stat_left, 1):
            _cell(ws, row, i, v if v else 0, STAT_NUM_FONT, WHITE_FILL, CENTER, THIN)
        self._sep_col(ws, row)
        # Right stat values
        for i, (_, v) in enumerate(stat_right, 6):
            _cell(ws, row, i, v if v else 0, STAT_NUM_FONT, WHITE_FILL, CENTER, THIN)
        _cell(ws, row, 10, total_records, STAT_NUM_FONT, ACCENT_FILL, CENTER, THIN)
        ws.row_dimensions[row].height = 26
        row += 2

        # --- LAST CALL / LOCATION ---
        self._section_header(ws, row, 'LAST CALL / LOCATION DETAILS')
        row += 1

        lc_headers = ['CALLED PARTY', 'CALL TYPE', 'DATE', 'TIME / DUR', '', 'IMEI', 'ROAMING CIRCLE', '', '']
        for i, h in enumerate(lc_headers, 1):
            if i == 5:
                _cell(ws, row, i, '', VALUE_FONT, SEP_FILL, border=None)
            elif h:
                _cell(ws, row, i, h, WHITE_SM, LIGHT_BLUE, CENTER)
        # Merge empty header cells
        ws.merge_cells(f'G{row}:J{row}')
        _cell(ws, row, 7, 'ROAMING CIRCLE', WHITE_SM, LIGHT_BLUE, CENTER)
        row += 1

        valid = df[df['CALL_DATETIME'].notna()].sort_values('CALL_DATETIME')
        if len(valid) > 0:
            last = valid.iloc[-1]
            party = self._safe(last.get('OTHER_PARTY', ''))
            op = self._safe(last.get('OP_LABEL', ''))
            circ = self._safe(last.get('ROAMING_CIRCLE', ''))
            party_display = party
            if op:
                party_display += f"\n({op})"

            dt = last.get('CALL_DATETIME', pd.NaT)
            date_s = dt.strftime('%d-%b-%Y') if pd.notna(dt) else ''
            time_s = dt.strftime('%H:%M:%S') if pd.notna(dt) else ''
            dur_s = self._fmt_dur(last.get('DUR_SECONDS', 0))
            time_dur = f"{time_s}\n{dur_s}" if dur_s else time_s

            _cell(ws, row, 1, party_display, VALUE_FONT, WHITE_FILL, LEFT_WRAP, THIN)
            _cell(ws, row, 2, self._safe(last.get('CALL_TYPE_LABEL', '')), VALUE_FONT, WHITE_FILL, CENTER, THIN)
            _cell(ws, row, 3, date_s, VALUE_FONT, WHITE_FILL, CENTER, THIN)
            _cell(ws, row, 4, time_dur, VALUE_FONT, WHITE_FILL, CENTER, THIN)
            self._sep_col(ws, row)
            _cell(ws, row, 6, self._safe(last.get('IMEI', '')), VALUE_FONT, WHITE_FILL, CENTER, THIN)
            ws.merge_cells(f'G{row}:J{row}')
            _cell(ws, row, 7, circ, VALUE_FONT, WHITE_FILL, CENTER, THIN)
            ws.row_dimensions[row].height = 30
            row += 1

            # Cell ID row - FIRST & LAST
            ws.merge_cells(f'A{row}:D{row}')
            _cell(ws, row, 1, 'FIRST CELL ID & ADDRESS', WHITE_SM, LIGHT_BLUE, CENTER)
            self._sep_col(ws, row)
            ws.merge_cells(f'F{row}:J{row}')
            _cell(ws, row, 6, 'LAST CELL ID & ADDRESS', WHITE_SM, LIGHT_BLUE, CENTER)
            row += 1

            first_cid = str(self._safe(last.get('FIRST_CELL_ID', '')))
            first_addr = str(self._safe(last.get('FIRST_CELL_ADDRESS', '')))
            last_cid = str(self._safe(last.get('LAST_CELL_ID', '')))
            last_addr = str(self._safe(last.get('LAST_CELL_ADDRESS', '')))
            first_cell = f"{first_cid}\n{first_addr}" if first_cid and first_addr else (first_cid or first_addr or 'N/A')
            last_cell = f"{last_cid}\n{last_addr}" if last_cid and last_addr else (last_cid or last_addr or 'N/A')

            ws.merge_cells(f'A{row}:D{row}')
            _cell(ws, row, 1, first_cell, VALUE_FONT, WHITE_FILL, LEFT_WRAP, THIN)
            self._sep_col(ws, row)
            ws.merge_cells(f'F{row}:J{row}')
            _cell(ws, row, 6, last_cell, VALUE_FONT, WHITE_FILL, LEFT_WRAP, THIN)
            ws.row_dimensions[row].height = 40
            row += 2

        # --- TOP 10 CALLERS ---
        self._section_header(ws, row, 'TOP 10 CALLERS')
        row += 1
        self._sub_header_pair(ws, row, 'MOST CALLS', 'MOST DURATION')
        row += 1

        # Left headers (Most Calls)
        left_h = ['CONTACT NUMBER', 'OPERATOR / CIRCLE', 'FIRST-LAST CALL', 'TOTAL CALLS']
        for i, h in enumerate(left_h, 1):
            _cell(ws, row, i, h, WHITE_SM, LIGHT_BLUE, CENTER)
        self._sep_col(ws, row)
        # Right headers (Most Duration)
        right_h = ['CONTACT NUMBER', 'OPERATOR / CIRCLE', 'FIRST-LAST CALL', 'DURATION', '']
        for i, h in enumerate(right_h, 6):
            if i <= 10:
                if h:
                    _cell(ws, row, i, h, WHITE_SM, LIGHT_BLUE, CENTER)
        # Merge last two right columns header
        ws.merge_cells(f'I{row}:J{row}')
        _cell(ws, row, 9, 'DURATION', WHITE_SM, LIGHT_BLUE, CENTER)
        row += 1

        # Get mobile-only contact stats
        if 'Contact Summary' in results and len(results['Contact Summary']) > 0:
            cs = results['Contact Summary']
            cs_mobile = cs[cs['TYPE'] == 'Mobile'] if 'TYPE' in cs.columns else cs
            top_calls = cs_mobile.head(10)
            top_dur = cs_mobile[cs_mobile['TOTAL DURATION'] > 0].sort_values('TOTAL DURATION', ascending=False).head(10)

            for idx in range(max(len(top_calls), len(top_dur))):
                fill = self._row_fill(idx)

                if idx < len(top_calls):
                    r = top_calls.iloc[idx]
                    fc = self._safe_date(r.get('FIRST CALL'))
                    lc = self._safe_date(r.get('LAST CALL'))
                    _cell(ws, row, 1, self._safe(r['CONTACT NUMBER']), VALUE_BOLD, fill, LEFT_WRAP, THIN)
                    op_circ = self._safe(r.get('TELECOM OPERATOR', ''))
                    circ_val = self._safe(r.get('CIRCLE', ''))
                    if circ_val:
                        op_circ += f"\n{circ_val}" if op_circ else circ_val
                    _cell(ws, row, 2, op_circ, VALUE_FONT, fill, LEFT_WRAP, THIN)
                    _cell(ws, row, 3, f"{fc}\n{lc}" if fc or lc else '', VALUE_FONT, fill, CENTER, THIN)
                    _cell(ws, row, 4, r['TOTAL CALLS'], STAT_NUM_FONT, fill, CENTER, THIN)
                else:
                    for c in range(1, 5):
                        _cell(ws, row, c, '', VALUE_FONT, fill, border=LIGHT_BORDER)

                self._sep_col(ws, row)

                if idx < len(top_dur):
                    r2 = top_dur.iloc[idx]
                    fc2 = self._safe_date(r2.get('FIRST CALL'))
                    lc2 = self._safe_date(r2.get('LAST CALL'))
                    _cell(ws, row, 6, self._safe(r2['CONTACT NUMBER']), VALUE_BOLD, fill, LEFT_WRAP, THIN)
                    op_circ2 = self._safe(r2.get('TELECOM OPERATOR', ''))
                    circ_val2 = self._safe(r2.get('CIRCLE', ''))
                    if circ_val2:
                        op_circ2 += f"\n{circ_val2}" if op_circ2 else circ_val2
                    _cell(ws, row, 7, op_circ2, VALUE_FONT, fill, LEFT_WRAP, THIN)
                    _cell(ws, row, 8, f"{fc2}\n{lc2}" if fc2 or lc2 else '', VALUE_FONT, fill, CENTER, THIN)
                    ws.merge_cells(f'I{row}:J{row}')
                    _cell(ws, row, 9, self._fmt_dur(r2['TOTAL DURATION']), STAT_NUM_FONT, fill, CENTER, THIN)
                else:
                    for c in range(6, 11):
                        _cell(ws, row, c, '', VALUE_FONT, fill, border=LIGHT_BORDER)

                ws.row_dimensions[row].height = 28
                row += 1
        row += 1

        # --- INTERNATIONAL CONTACTS ---
        self._section_header(ws, row, 'INTERNATIONAL CONTACTS')
        row += 1
        self._sub_header_pair(ws, row, 'MOST CALLS', 'MOST DURATION')
        row += 1

        for i, h in enumerate(['CONTACT NUMBER', 'COUNTRY', 'CALL TYPE', 'TOTAL CALLS'], 1):
            _cell(ws, row, i, h, WHITE_SM, LIGHT_BLUE, CENTER)
        self._sep_col(ws, row)
        for i, h in enumerate(['CONTACT NUMBER', 'COUNTRY', 'CALL TYPE', 'DURATION', ''], 6):
            if i <= 10 and h:
                _cell(ws, row, i, h, WHITE_SM, LIGHT_BLUE, CENTER)
        ws.merge_cells(f'I{row}:J{row}')
        _cell(ws, row, 9, 'DURATION', WHITE_SM, LIGHT_BLUE, CENTER)
        row += 1

        if 'Contact Summary' in results:
            isd = results['Contact Summary']
            isd = isd[isd['TYPE'] == 'ISD'] if 'TYPE' in isd.columns else pd.DataFrame()
            isd_by_calls = isd.head(5) if len(isd) > 0 else pd.DataFrame()
            isd_by_dur = isd[isd['TOTAL DURATION'] > 0].sort_values('TOTAL DURATION', ascending=False).head(5) if len(isd) > 0 else pd.DataFrame()

            if len(isd) == 0:
                ws.merge_cells(f'A{row}:J{row}')
                _cell(ws, row, 1, 'No International Contacts Found', VALUE_FONT, WHITE_FILL, CENTER, THIN)
                row += 1
            else:
                for idx in range(max(len(isd_by_calls), len(isd_by_dur))):
                    if idx < len(isd_by_calls):
                        r = isd_by_calls.iloc[idx]
                        _cell(ws, row, 1, self._safe(r['CONTACT NUMBER']), VALUE_FONT, WHITE_FILL, LEFT_WRAP, THIN)
                        _cell(ws, row, 2, '', VALUE_FONT, WHITE_FILL, CENTER, THIN)
                        _cell(ws, row, 3, '', VALUE_FONT, WHITE_FILL, CENTER, THIN)
                        _cell(ws, row, 4, r['TOTAL CALLS'], VALUE_FONT, WHITE_FILL, CENTER, THIN)

                    self._sep_col(ws, row)

                    if idx < len(isd_by_dur):
                        r2 = isd_by_dur.iloc[idx]
                        _cell(ws, row, 6, self._safe(r2['CONTACT NUMBER']), VALUE_FONT, WHITE_FILL, LEFT_WRAP, THIN)
                        _cell(ws, row, 7, '', VALUE_FONT, WHITE_FILL, CENTER, THIN)
                        _cell(ws, row, 8, '', VALUE_FONT, WHITE_FILL, CENTER, THIN)
                        ws.merge_cells(f'I{row}:J{row}')
                        _cell(ws, row, 9, self._fmt_dur(r2['TOTAL DURATION']), VALUE_FONT, WHITE_FILL, CENTER, THIN)
                    row += 1
        row += 1

        # --- IMEI SUMMARY ---
        self._section_header(ws, row, 'IMEI SUMMARY')
        row += 1

        # Headers - IMEI spans 2 cols, MODEL spans 3 cols
        ws.merge_cells(f'A{row}:B{row}')
        _cell(ws, row, 1, 'IMEI NUMBER', WHITE_SM, LIGHT_BLUE, CENTER)
        ws.merge_cells(f'C{row}:D{row}')
        _cell(ws, row, 3, 'HANDSET MODEL', WHITE_SM, LIGHT_BLUE, CENTER)
        _cell(ws, row, 5, '', VALUE_FONT, SEP_FILL, border=None)
        for i, h in enumerate(['OUT-CALL', 'IN-CALL', 'SMS', 'FIRST USE', 'LAST USE'], 6):
            _cell(ws, row, i, h, WHITE_SM, LIGHT_BLUE, CENTER)
        row += 1

        if 'IMEI Summary' in results and len(results['IMEI Summary']) > 0:
            for imei_idx, (_, r) in enumerate(results['IMEI Summary'].iterrows()):
                fill = self._row_fill(imei_idx)
                ws.merge_cells(f'A{row}:B{row}')
                _cell(ws, row, 1, self._safe(r.get('IMEI', '')), VALUE_BOLD, fill, LEFT_WRAP, THIN)
                ws.merge_cells(f'C{row}:D{row}')
                _cell(ws, row, 3, self._safe(r.get('HANDSET', '')), VALUE_FONT, fill, LEFT_WRAP, THIN)
                self._sep_col(ws, row)

                out_c = r.get('OUT-CALL', '')
                in_c = r.get('IN-CALL', '')
                sms_total = 0
                try:
                    sms_total = int(r.get('IN-SMS', 0) or 0) + int(r.get('OUT-SMS', 0) or 0) + int(r.get('BSM', 0) or 0)
                except (ValueError, TypeError):
                    pass

                _cell(ws, row, 6, out_c if out_c else '', VALUE_FONT, fill, CENTER, THIN)
                _cell(ws, row, 7, in_c if in_c else '', VALUE_FONT, fill, CENTER, THIN)
                _cell(ws, row, 8, sms_total if sms_total else '', VALUE_FONT, fill, CENTER, THIN)

                fc = r.get('FIRST CALL')
                lc = r.get('LAST CALL')
                _cell(ws, row, 9, self._safe_date(fc), VALUE_FONT, fill, CENTER, THIN)
                _cell(ws, row, 10, self._safe_date(lc), VALUE_FONT, fill, CENTER, THIN)
                ws.row_dimensions[row].height = 20
                row += 1
        else:
            ws.merge_cells(f'A{row}:J{row}')
            _cell(ws, row, 1, 'No IMEI Data Available', VALUE_FONT, WHITE_FILL, CENTER, THIN)
            row += 1
        row += 1

        # --- TOP 10 LOCATIONS ---
        self._section_header(ws, row, 'TOP 10 LOCATIONS')
        row += 1

        ws.merge_cells(f'A{row}:B{row}')
        _cell(ws, row, 1, 'CELL-ID', WHITE_SM, LIGHT_BLUE, CENTER)
        ws.merge_cells(f'C{row}:G{row}')
        _cell(ws, row, 3, 'CELL ADDRESS / LOCATION', WHITE_SM, LIGHT_BLUE, CENTER)
        _cell(ws, row, 8, 'FIRST CALL', WHITE_SM, LIGHT_BLUE, CENTER)
        _cell(ws, row, 9, 'LAST CALL', WHITE_SM, LIGHT_BLUE, CENTER)
        _cell(ws, row, 10, 'TOTAL', WHITE_SM, LIGHT_BLUE, CENTER)
        row += 1

        if 'Location Summary' in results and len(results['Location Summary']) > 0:
            for loc_idx, (_, r) in enumerate(results['Location Summary'].head(10).iterrows()):
                fill = self._row_fill(loc_idx)
                fc = r.get('FIRST CALL')
                lc = r.get('LAST CALL')
                fc_s = self._safe_date(fc)
                lc_s = self._safe_date(lc)

                ws.merge_cells(f'A{row}:B{row}')
                _cell(ws, row, 1, self._safe(r.get('CELL ID', '')), VALUE_FONT, fill, LEFT_WRAP, THIN)
                ws.merge_cells(f'C{row}:G{row}')
                addr = str(self._safe(r.get('CELL ADDRESS', r.get('ADDRESS', ''))))[:150]
                _cell(ws, row, 3, addr, VALUE_FONT, fill, LEFT_WRAP, THIN)
                _cell(ws, row, 8, fc_s, VALUE_FONT, fill, CENTER, THIN)
                _cell(ws, row, 9, lc_s, VALUE_FONT, fill, CENTER, THIN)
                _cell(ws, row, 10, r.get('TOTAL CALLS', 0), STAT_NUM_FONT, fill, CENTER, THIN)
                ws.row_dimensions[row].height = 30
                row += 1
        else:
            ws.merge_cells(f'A{row}:J{row}')
            _cell(ws, row, 1, 'No Location Data Available', VALUE_FONT, WHITE_FILL, CENTER, THIN)
            row += 1
        row += 1

        # --- ROAMING SUMMARY (if data exists) ---
        if 'Roaming Summary' in results and len(results['Roaming Summary']) > 0:
            self._section_header(ws, row, 'ROAMING SUMMARY')
            row += 1

            _cell(ws, row, 1, 'SR.', WHITE_SM, LIGHT_BLUE, CENTER)
            ws.merge_cells(f'B{row}:D{row}')
            _cell(ws, row, 2, 'ROAMING CIRCLE', WHITE_SM, LIGHT_BLUE, CENTER)
            self._sep_col(ws, row)
            ws.merge_cells(f'F{row}:H{row}')
            _cell(ws, row, 6, 'FROM DATE-TIME', WHITE_SM, LIGHT_BLUE, CENTER)
            ws.merge_cells(f'I{row}:J{row}')
            _cell(ws, row, 9, 'TO DATE-TIME', WHITE_SM, LIGHT_BLUE, CENTER)
            row += 1

            for roam_idx, (_, r) in enumerate(results['Roaming Summary'].iterrows()):
                fill = self._row_fill(roam_idx)
                _cell(ws, row, 1, roam_idx + 1, SR_FONT, fill, CENTER, THIN)
                ws.merge_cells(f'B{row}:D{row}')
                _cell(ws, row, 2, self._safe(r.get('ROAMING CIRCLE', '')), VALUE_BOLD, fill, CENTER, THIN)
                self._sep_col(ws, row)
                fr = r.get('FROM DATE-TIME')
                to = r.get('TO DATE-TIME')
                ws.merge_cells(f'F{row}:H{row}')
                _cell(ws, row, 6, self._safe_date(fr, '%d-%b-%Y %H:%M'), VALUE_FONT, fill, CENTER, THIN)
                ws.merge_cells(f'I{row}:J{row}')
                _cell(ws, row, 9, self._safe_date(to, '%d-%b-%Y %H:%M'), VALUE_FONT, fill, CENTER, THIN)
                ws.row_dimensions[row].height = 18
                row += 1
            row += 1

        # --- FOOTER ---
        row += 1  # extra spacing before footer
        ws.merge_cells(f'A{row}:J{row}')
        _cell(ws, row, 1, f'Report Generated by CDR Forensic Analysis Tool v3.0 | Target: {target}',
              FOOTER_FONT, border=None, align=CENTER)

        # --- COLUMN WIDTHS ---
        col_widths = {'A': 22, 'B': 18, 'C': 16, 'D': 14, 'E': 2.5, 'F': 22, 'G': 18, 'H': 16, 'I': 15, 'J': 15}
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w

        # Print settings
        ws.sheet_properties.pageSetUpPr = None
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.print_area = f'A1:J{row}'

    # ==============================================================
    # RAW CDR
    # ==============================================================
    def _add_raw_cdr(self, wb, metadata):
        """Add raw file content as last sheet"""
        ws = wb.create_sheet('Raw CDR')
        csv_path = metadata.get('_csv_path', '')
        if not csv_path:
            return
        try:
            # Try multiple encodings
            enc = 'utf-8'
            for try_enc in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
                try:
                    with open(csv_path, 'r', encoding=try_enc) as f:
                        f.read(1024)
                    enc = try_enc
                    break
                except (UnicodeDecodeError, UnicodeError):
                    continue

            with open(csv_path, 'r', encoding=enc) as f:
                for ridx, line in enumerate(f, 1):
                    if ridx > 5000:
                        break
                    parts = line.strip().split(',')
                    for cidx, val in enumerate(parts[:22], 1):
                        _cell(ws, ridx, cidx, val.strip("'\""), VALUE_FONT, border=None)
        except Exception as e:
            logger.warning(f"  Could not write Raw CDR: {e}")

    # ==============================================================
    # FORMATTING
    # ==============================================================
    def _format_all(self, wb):
        """Apply consistent formatting to all sheets"""
        from openpyxl.utils import get_column_letter

        # Sheet tab colors
        tab_colors = {
            'Smart Report': '1F4E79',
            'FULL CDR - PRINT': '2E75B6',
            'Full CDR': '2E75B6',
            'Contact Summary': '548235',
            'Contact Duration Summary': '548235',
            'Top Callers': '548235',
            'Top Called': '548235',
            'Roaming Contact Summary': 'BF8F00',
            'IMEI Summary': '843C0C',
            'Roaming Summary': 'BF8F00',
            'Location Summary': '7030A0',
            'Festival Calls': 'C00000',
            'Hourly Activity': '7030A0',
            'Weekday Analysis': '7030A0',
            'Common IMEI': '843C0C',
            'Conference or Call Hold': 'C00000',
            'IMSI Summary': '843C0C',
            'Rejected Data': '808080',
        }

        for sn in wb.sheetnames:
            ws = wb[sn]

            # Apply tab colors
            if sn in tab_colors:
                ws.sheet_properties.tabColor = tab_colors[sn]
            elif 'Halt' in sn:
                ws.sheet_properties.tabColor = '4472C4'
            elif sn == 'Raw CDR':
                ws.sheet_properties.tabColor = '808080'

            if sn == 'Smart Report':
                continue

            # Auto column widths
            for col_idx in range(1, min(ws.max_column + 1, 22)):
                letter = get_column_letter(col_idx)
                max_len = 12
                for r in range(1, min(ws.max_row + 1, 50)):
                    v = ws.cell(r, col_idx).value
                    if v:
                        val_str = str(v).split('\n')[0]
                        # Filter nan from display
                        if val_str.lower() not in ('nan', 'nat', 'none'):
                            max_len = max(max_len, min(len(val_str), 50))
                ws.column_dimensions[letter].width = max_len + 2

    # ==============================================================
    # HELPERS
    # ==============================================================
    @staticmethod
    def _sheet_title(name):
        titles = {
            'FULL CDR - PRINT': 'Full CDR (Print)',
            'Full CDR': 'CDR Report',
            'Contact Summary': 'Max Caller',
            'Contact Duration Summary': 'Max Duration',
            'Top Callers': 'Top Callers (Incoming)',
            'Top Called': 'Top Called (Outgoing)',
            'Roaming Contact Summary': 'Roaming Contact Summary',
            'IMEI Summary': 'IMEI Summary',
            'Roaming Summary': 'Roaming Circle',
            'Location Summary': 'Max Location',
            'Festival Calls': 'Festival Calls',
            'Day Halt': 'Day Halt',
            'Day Halt Caller': 'Day Halt Caller',
            'Day Halt CDR': 'Day Halt CDR',
            'Night Halt': 'Night Halt',
            'Night Halt Caller': 'Night Halt Caller',
            'Night Halt CDR': 'Night Halt CDR',
            'Conference or Call Hold': 'Conference / Call Hold Details',
            'IMSI Summary': 'IMSI Summary',
            'Rejected Data': 'Rejected Records',
            'Hourly Activity': 'Hourly Activity Analysis',
            'Weekday Analysis': 'Weekday Activity Analysis',
            'Common IMEI': 'Common IMEI (Device Sharing)',
        }
        return titles.get(name, name)

    @staticmethod
    def _format_value(val):
        try:
            if val is None:
                return ''
            if isinstance(val, pd.Timestamp):
                return val.strftime('%d-%b-%Y %H:%M:%S') if pd.notna(val) else ''
            try:
                if pd.isna(val):
                    return ''
            except (TypeError, ValueError):
                pass
            s = str(val).strip()
            if s.lower() in ['nan', 'nat', 'none', '']:
                return ''
            return val
        except Exception:
            return str(val) if val is not None else '' 
